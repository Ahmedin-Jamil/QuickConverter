"""
Load Layer - Multi-sheet Excel generation with professional formatting.

Generates interview-grade Excel output with:
1. Transactions sheet - Full transaction data with categories
2. Financial Summary sheet - Totals, balances, and reconciliation check
3. Data Quality Report sheet - Clean/flagged rows with reasons
4. Audit Trail sheet - Processing metadata
"""
import pandas as pd
from io import BytesIO
from typing import List, Dict, Any
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class UniversalLoader:
    """
    Universal exporter for multiple document formats.
    Supported: 'xlsx', 'csv', 'txt'
    """
    
    def __init__(self):
        self.currency_format = '$#,##0.00'
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        self.success_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        self.border = Border(bottom=Side(style='thin', color='DDDDDD'))

    def generate(self, transactions: List[Dict], audit_data: Dict[str, Any], target_format: str = "xlsx") -> BytesIO:
        """
        Generate output in the requested format.
        """
        if target_format == "csv":
            return self._generate_csv(transactions)
        elif target_format == "txt":
            return self._generate_text(transactions, audit_data)
        else:
            return self._generate_excel(transactions, audit_data)

    def _generate_excel(self, transactions: List[Dict], audit_data: Dict[str, Any]) -> BytesIO:
        """
        Structured Excel Report with multiple sheets:
        1. Transactions - Full data with categories
        2. Financial Summary - Totals and reconciliation check
        3. Data Quality Report - Flagged rows with reasons
        """
        output = BytesIO()
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # ════════════════════════════════════════════════════════════════
        # SHEET 1: TRANSACTIONS
        # ════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Transactions"
        
        headers = ["Date", "Description", "Category", "Debit", "Credit", "Balance", "DQ Flag"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, tx in enumerate(transactions, 2):
            tx_type = tx.get("tx_type", "debit")
            amount = tx.get("amount", 0.0)
            dq_flag = tx.get("metadata", {}).get("dq_flag", "unknown")
            
            row_data = [
                tx.get("post_date"),
                tx.get("description"),
                tx.get("category", "Uncategorized"),
                amount if tx_type == "debit" else None,
                amount if tx_type == "credit" else None,
                tx.get("balance"),
                dq_flag.replace('_', ' ').upper()
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)
                if col_idx in [4, 5, 6]:  # Debit, Credit, Balance
                    cell.number_format = self.currency_format
                cell.border = self.border
        
        self._auto_width(ws1)
        ws1.freeze_panes = "A2"
        
        # ════════════════════════════════════════════════════════════════
        # SHEET 2: FINANCIAL SUMMARY
        # ════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Financial Summary")
        
        financials = audit_data.get("financials", {})
        reconciliation = audit_data.get("reconciliation", {})
        
        total_debits = financials.get("total_debits", 0)
        total_credits = financials.get("total_credits", 0)
        opening = financials.get("opening_balance", 0)
        closing = financials.get("closing_balance", 0)
        
        # Title
        ws2.cell(row=1, column=1, value="FINANCIAL SUMMARY").font = Font(bold=True, size=14)
        ws2.merge_cells('A1:B1')
        
        summary_items = [
            ("", ""),  # Spacer
            ("Opening Balance", opening),
            ("Total Credits (+)", total_credits),
            ("Total Debits (-)", total_debits),
            ("Net Change", total_credits - total_debits),
            ("Closing Balance", closing),
        ]
        
        row = 3
        for key, val in summary_items:
            if key:
                ws2.cell(row=row, column=1, value=key).font = Font(bold=True)
                c = ws2.cell(row=row, column=2, value=val)
                if isinstance(val, (int, float)):
                    c.number_format = self.currency_format
            row += 1
        
        # Reconciliation Check Section
        row += 1
        ws2.cell(row=row, column=1, value="RECONCILIATION CHECK").font = Font(bold=True, size=12)
        row += 1
        
        expected = reconciliation.get("expected_closing", 0)
        actual = reconciliation.get("actual_closing", 0)
        is_balanced = reconciliation.get("is_balanced", True)
        status = reconciliation.get("status", "N/A")
        
        recon_items = [
            ("Opening + Credits - Debits =", expected),
            ("Actual Closing Balance =", actual),
            ("Status", status),
        ]
        
        for key, val in recon_items:
            ws2.cell(row=row, column=1, value=key).font = Font(bold=True)
            c = ws2.cell(row=row, column=2, value=val)
            if isinstance(val, (int, float)):
                c.number_format = self.currency_format
            if key == "Status":
                c.fill = self.success_fill if is_balanced else self.warning_fill
                c.font = Font(bold=True)
            row += 1
        
        self._auto_width(ws2)
        
        # ════════════════════════════════════════════════════════════════
        # SHEET 3: DATA QUALITY REPORT
        # ════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Data Quality Report")
        
        dq_report = audit_data.get("dq_report", {})
        dq_stats = dq_report.get("stats", audit_data.get("dq_stats", {}))
        flagged_rows = dq_report.get("flagged_rows", [])
        
        # Title
        ws3.cell(row=1, column=1, value="DATA QUALITY REPORT").font = Font(bold=True, size=14)
        ws3.merge_cells('A1:D1')
        
        # Summary Statistics
        row = 3
        ws3.cell(row=row, column=1, value="Summary Statistics").font = Font(bold=True, size=12)
        row += 1
        
        stat_items = [
            ("Total Transactions", dq_stats.get("total", 0)),
            ("Clean Transactions (Table)", dq_stats.get("CLEAN", 0)),
            ("Recovered Transactions (Regex)", dq_stats.get("RECOVERED_TRANSACTION", 0)),
            ("Suspect Rows (Anomalies)", dq_stats.get("SUSPECT", 0)),
            ("Non-Transaction Rows (Metadata)", dq_stats.get("NON_TRANSACTION", 0)),
            ("Total Flags", len(flagged_rows)),
        ]
        
        for key, val in stat_items:
            ws3.cell(row=row, column=1, value=key)
            ws3.cell(row=row, column=2, value=val)
            row += 1
        
        # Flagged Rows Table
        row += 1
        ws3.cell(row=row, column=1, value="Flagged Rows Detail").font = Font(bold=True, size=12)
        row += 1
        
        if flagged_rows:
            flag_headers = ["Row #", "Date", "Description", "Amount", "Flag Type", "Reason"]
            for col_idx, header in enumerate(flag_headers, 1):
                cell = ws3.cell(row=row, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            row += 1
            
            for flag in flagged_rows:
                ws3.cell(row=row, column=1, value=flag.get("row", ""))
                ws3.cell(row=row, column=2, value=flag.get("date", ""))
                ws3.cell(row=row, column=3, value=flag.get("description", ""))
                c = ws3.cell(row=row, column=4, value=flag.get("amount", 0))
                c.number_format = self.currency_format
                ws3.cell(row=row, column=5, value=flag.get("flag_type", "").upper())
                ws3.cell(row=row, column=6, value=flag.get("reason", ""))
                row += 1
        else:
            ws3.cell(row=row, column=1, value="✅ No flagged rows - all data passed quality checks").fill = self.success_fill
        
        self._auto_width(ws3)
        
        wb.save(output)
        output.seek(0)
        return output

    def _generate_csv(self, transactions: List[Dict]) -> BytesIO:
        """Simple CSV export for interoperability - includes category"""
        flattened = []
        for tx in transactions:
            tx_type = tx.get("tx_type", "debit")
            amount = tx.get("amount", 0.0)
            flattened.append({
                "Date": tx.get("post_date"),
                "Description": tx.get("description"),
                "Category": tx.get("category", "Uncategorized"),
                "Debit": amount if tx_type == "debit" else 0.0,
                "Credit": amount if tx_type == "credit" else 0.0,
                "Balance": tx.get("balance") or 0.0,
                "DQ_Flag": tx.get("metadata", {}).get("dq_flag", "unknown")
            })
        
        df = pd.DataFrame(flattened)
        output = BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        return output

    def _generate_text(self, transactions: List[Dict], audit_data: Dict[str, Any]) -> BytesIO:
        """Structured text export for debugging/preview"""
        output = BytesIO()
        lines = [f"QCONVERTER DOCUMENT REPORT - {audit_data.get('timestamp')}\n", "="*50 + "\n\n"]
        
        for tx in transactions:
            line = f"[{tx.get('post_date')}] {tx.get('description')[:40]:<40} | Amt: {tx.get('amount'):>10.2f} | Type: {tx.get('tx_type')} | Cat: {tx.get('category', 'N/A')}\n"
            lines.append(line)
            
        output.write("".join(lines).encode('utf-8'))
        output.seek(0)
        return output

    def _auto_width(self, ws) -> None:
        """Auto-adjust column widths"""
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 60)

